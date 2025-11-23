"""
Utilidades de formato para exportaciones Excel
Define colores, estilos y formateo según el diseño del sistema
"""
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelFormatters:
    """Clase con estilos y formateo reutilizables para Excel"""
    
    # Colores del sistema
    PRIMARY_COLOR = "3b82f6"  # Azul primary
    PRIMARY_DARK = "2563eb"   # Azul primary oscuro
    SUCCESS_COLOR = "10b981"  # Verde success
    TEAL_COLOR = "56c3a8"     # Verde-azul (teal) del gradiente
    BLUE_COLOR = "4cabff"     # Azul del gradiente
    GRAY_LIGHT = "f8fafc"     # Gris muy claro para filas alternadas
    GRAY_MEDIUM = "e5e7eb"    # Gris medio
    GRAY_BORDER = "e2e8f0"    # Gris para bordes
    GRAY_TEXT = "6b7280"      # Gris para texto secundario
    WHITE = "FFFFFF"           # Blanco
    
    @staticmethod
    def get_header_style():
        """Estilo para encabezados de tabla"""
        return {
            'fill': PatternFill(
                start_color=ExcelFormatters.PRIMARY_COLOR,
                end_color=ExcelFormatters.PRIMARY_COLOR,
                fill_type="solid"
            ),
            'font': Font(
                bold=True,
                color=ExcelFormatters.WHITE,
                size=11,
                name="Arial"
            ),
            'alignment': Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            ),
            'border': Border(
                left=Side(style='thin', color=ExcelFormatters.GRAY_BORDER),
                right=Side(style='thin', color=ExcelFormatters.GRAY_BORDER),
                top=Side(style='thin', color=ExcelFormatters.GRAY_BORDER),
                bottom=Side(style='thin', color=ExcelFormatters.GRAY_BORDER)
            )
        }
    
    @staticmethod
    def get_data_style(alternate=False):
        """Estilo para datos de tabla"""
        fill_color = ExcelFormatters.GRAY_LIGHT if alternate else ExcelFormatters.WHITE
        
        return {
            'fill': PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type="solid"
            ),
            'font': Font(
                size=11,
                name="Arial"
            ),
            'alignment': Alignment(
                vertical="center",
                wrap_text=True
            ),
            'border': Border(
                left=Side(style='thin', color=ExcelFormatters.GRAY_BORDER),
                right=Side(style='thin', color=ExcelFormatters.GRAY_BORDER),
                top=Side(style='thin', color=ExcelFormatters.GRAY_BORDER),
                bottom=Side(style='thin', color=ExcelFormatters.GRAY_BORDER)
            )
        }
    
    @staticmethod
    def format_date(date_value):
        """Formatea una fecha para Excel"""
        if date_value is None:
            return ""
        if hasattr(date_value, 'strftime'):
            return date_value.strftime('%Y-%m-%d %H:%M:%S')
        return str(date_value)
    
    @staticmethod
    def format_text(text_value):
        """Formatea texto, maneja None"""
        if text_value is None:
            return ""
        return str(text_value)
    
    @staticmethod
    def get_column_widths(headers):
        """Calcula anchos de columna automáticos basados en headers"""
        widths = {}
        for idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(idx)
            # Ancho mínimo 12, máximo 50, basado en longitud del header + margen
            width = min(max(len(str(header)) + 4, 12), 50)
            widths[col_letter] = width
        return widths
    
    @staticmethod
    def get_title_style():
        """Estilo para el título principal del reporte"""
        return {
            'fill': PatternFill(
                start_color=ExcelFormatters.PRIMARY_COLOR,
                end_color=ExcelFormatters.PRIMARY_DARK,
                fill_type="solid"
            ),
            'font': Font(
                bold=True,
                color=ExcelFormatters.WHITE,
                size=16,
                name="Arial"
            ),
            'alignment': Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=False
            ),
            'border': Border(
                left=Side(style='thin', color=ExcelFormatters.PRIMARY_DARK),
                right=Side(style='thin', color=ExcelFormatters.PRIMARY_DARK),
                top=Side(style='thin', color=ExcelFormatters.PRIMARY_DARK),
                bottom=Side(style='medium', color=ExcelFormatters.PRIMARY_DARK)
            )
        }
    
    @staticmethod
    def get_info_style():
        """Estilo para información secundaria (fechas, entidad, etc.)"""
        return {
            'fill': PatternFill(
                start_color=ExcelFormatters.GRAY_LIGHT,
                end_color=ExcelFormatters.GRAY_LIGHT,
                fill_type="solid"
            ),
            'font': Font(
                bold=False,
                color=ExcelFormatters.GRAY_TEXT,
                size=10,
                name="Arial"
            ),
            'alignment': Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
        }
    
    @staticmethod
    def get_label_style():
        """Estilo para etiquetas de información"""
        return {
            'font': Font(
                bold=True,
                color=ExcelFormatters.PRIMARY_COLOR,
                size=10,
                name="Arial"
            ),
            'alignment': Alignment(
                horizontal="left",
                vertical="center"
            )
        }

