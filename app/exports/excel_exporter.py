"""
Exportador de Excel con formato profesional
Genera archivos Excel con diseño del sistema aplicado
"""
from io import BytesIO
from typing import List, Dict, Any, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from app.exports.formatters import ExcelFormatters


class ExcelExporter:
    """Clase para generar archivos Excel con formato profesional"""
    
    # Diccionario de traducción de entidades al español
    ENTITY_TRANSLATIONS = {
        'user': 'Usuarios',
        'users': 'Usuarios',
        'User': 'Usuarios',
        'Users': 'Usuarios',
        'employee': 'Empleados',
        'employees': 'Empleados',
        'Employee': 'Empleados',
        'Employees': 'Empleados',
        'tank': 'Tanques',
        'tanks': 'Tanques',
        'Tank': 'Tanques',
        'Tanks': 'Tanques',
        'pipe': 'Tuberías',
        'pipes': 'Tuberías',
        'Pipe': 'Tuberías',
        'Pipes': 'Tuberías',
        'connection': 'Conexiones',
        'connections': 'Conexiones',
        'Connection': 'Conexiones',
        'Connections': 'Conexiones',
        'intervention': 'Intervenciones',
        'interventions': 'Intervenciones',
        'Intervention': 'Intervenciones',
        'Interventions': 'Intervenciones',
        'rol': 'Roles',
        'roles': 'Roles',
        'Rol': 'Roles',
        'Roles': 'Roles',
        'permission': 'Permisos',
        'permissions': 'Permisos',
        'Permission': 'Permisos',
        'Permissions': 'Permisos',
        'type_employee': 'Tipos de Empleado',
        'type_employees': 'Tipos de Empleado',
        'TypeEmployee': 'Tipos de Empleado',
    }
    
    def __init__(self):
        self.formatters = ExcelFormatters()
    
    def translate_entity(self, entity_name: str) -> str:
        """Traduce el nombre de la entidad al español"""
        return self.ENTITY_TRANSLATIONS.get(entity_name, entity_name.capitalize())
    
    def create_workbook(self) -> Workbook:
        """Crea un nuevo workbook"""
        return Workbook()
    
    def apply_header_style(self, worksheet, header_row: int = 1):
        """Aplica estilo a la fila de encabezados"""
        header_style = self.formatters.get_header_style()
        
        for cell in worksheet[header_row]:
            for attr, value in header_style.items():
                setattr(cell, attr, value)
        
        # Ajustar altura de fila de encabezados
        worksheet.row_dimensions[header_row].height = 25
    
    def apply_data_formatting(self, worksheet, start_row: int = 2, end_row: int = None):
        """Aplica formato a las filas de datos"""
        if end_row is None:
            end_row = worksheet.max_row
        
        for row_num in range(start_row, end_row + 1):
            alternate = (row_num - start_row) % 2 == 0
            data_style = self.formatters.get_data_style(alternate=alternate)
            
            for cell in worksheet[row_num]:
                for attr, value in data_style.items():
                    setattr(cell, attr, value)
            
            # Ajustar altura de fila
            worksheet.row_dimensions[row_num].height = 18
    
    def auto_adjust_columns(self, worksheet, headers: List[str]):
        """Ajusta automáticamente el ancho de las columnas"""
        # Obtener anchos base desde formatters
        base_widths = self.formatters.get_column_widths(headers)
        
        # Ajustar según contenido real
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_length = len(str(header))
            
            # Revisar todas las celdas de la columna para encontrar el contenido más largo
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
            
            # Aplicar ancho con márgenes
            width = min(max(max_length + 4, 12), 50)
            worksheet.column_dimensions[col_letter].width = width
    
    def add_report_header(
        self,
        worksheet,
        entity_name: str,
        date_start: Optional[str] = None,
        date_end: Optional[str] = None,
        user_name: Optional[str] = None
    ):
        """Agrega un encabezado elegante al reporte, centrado como en la referencia"""
        # Traducir nombre de entidad al español
        entity_spanish = self.translate_entity(entity_name)
        
        # Título principal - Centrado en columnas B-G (como en la referencia)
        title_cell = worksheet['B1']
        title_cell.value = f"REPORTE DE LOGS - {entity_spanish.upper()}"
        title_style = self.formatters.get_title_style()
        for attr, value in title_style.items():
            setattr(title_cell, attr, value)
        
        # Combinar celdas para el título (B1:G1) - centrado como en la referencia
        worksheet.merge_cells('B1:G1')
        worksheet.row_dimensions[1].height = 35
        
        # Información del reporte - Centrada también
        current_row = 3
        
        # Fecha de generación - Centrada
        gen_date_cell = worksheet[f'B{current_row}']
        gen_date_cell.value = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        info_style = self.formatters.get_info_style()
        # Centrar información
        info_style['alignment'] = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for attr, value in info_style.items():
            setattr(gen_date_cell, attr, value)
        worksheet.merge_cells(f'B{current_row}:G{current_row}')
        worksheet.row_dimensions[current_row].height = 20
        current_row += 1
        
        # Rango de fechas - Centrado
        if date_start and date_end:
            try:
                # Formatear fechas para mostrar
                start_date = datetime.fromisoformat(date_start.replace('Z', '+00:00'))
                end_date = datetime.fromisoformat(date_end.replace('Z', '+00:00'))
                date_range_cell = worksheet[f'B{current_row}']
                date_range_cell.value = f"Período: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
                for attr, value in info_style.items():
                    setattr(date_range_cell, attr, value)
                worksheet.merge_cells(f'B{current_row}:G{current_row}')
                worksheet.row_dimensions[current_row].height = 20
                current_row += 1
            except:
                pass
        
        # Usuario que genera el reporte - Centrado
        if user_name:
            user_cell = worksheet[f'B{current_row}']
            user_cell.value = f"Generado por: {user_name}"
            for attr, value in info_style.items():
                setattr(user_cell, attr, value)
            worksheet.merge_cells(f'B{current_row}:G{current_row}')
            worksheet.row_dimensions[current_row].height = 20
            current_row += 1
        
        # Línea separadora - Centrada también
        current_row += 1
        separator_cell = worksheet[f'B{current_row}']
        separator_cell.value = ""
        separator_fill = PatternFill(
            start_color=ExcelFormatters.GRAY_MEDIUM,
            end_color=ExcelFormatters.GRAY_MEDIUM,
            fill_type="solid"
        )
        separator_cell.fill = separator_fill
        worksheet.merge_cells(f'B{current_row}:G{current_row}')
        worksheet.row_dimensions[current_row].height = 3
        
        return current_row + 1  # Retornar la siguiente fila para los datos
    
    def export_logs(
        self,
        logs: List[Any],
        filters: Dict[str, Any] = None,
        user_info: Dict[str, Any] = None
    ) -> BytesIO:
        """
        Exporta logs a Excel con formato profesional
        
        Args:
            logs: Lista de objetos Logs
            filters: Diccionario con filtros aplicados
            user_info: Información del usuario que genera el reporte
        
        Returns:
            BytesIO: Buffer con el archivo Excel
        """
        wb = self.create_workbook()
        ws = wb.active
        ws.title = "Reportes"
        
        # Obtener información para el encabezado
        entity_name = filters.get('name_entity', 'Sistema') if filters else 'Sistema'
        date_start = filters.get('date_start') if filters else None
        date_end = filters.get('date_finish') if filters else None
        user_name = user_info.get('username', 'Usuario') if user_info else 'Usuario'
        
        # Agregar encabezado del reporte
        data_start_row = self.add_report_header(
            ws,
            entity_name=entity_name,
            date_start=date_start,
            date_end=date_end,
            user_name=user_name
        )
        
        # Definir encabezados de tabla (empezando desde columna A)
        headers = [
            'ID',
            'Fecha',
            'Usuario',
            'Acción',
            'Entidad',
            'ID Entidad',
            'Descripción'
        ]
        
        # Agregar encabezados de tabla (desde columna A)
        header_row = data_start_row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
        
        self.apply_header_style(ws, header_row=header_row)
        
        # Agregar datos
        data_row = header_row + 1
        for log in logs:
            # Obtener nombre de usuario si existe relación
            username = ""
            if hasattr(log, 'user') and log.user:
                username = getattr(log.user, 'user', '')
            
            row_data = [
                log.log_id,
                self.formatters.format_date(log.created_at),
                username,
                self.formatters.format_text(log.action),
                self.formatters.format_text(log.entity),
                log.entity_id if log.entity_id else "",
                self.formatters.format_text(log.description)
            ]
            
            # Agregar datos fila por fila
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=data_row, column=col_idx)
                cell.value = value
            
            data_row += 1
        
        # Aplicar formato a datos
        if data_row > header_row + 1:
            self.apply_data_formatting(ws, start_row=header_row + 1, end_row=data_row - 1)
        
        # Ajustar columnas automáticamente
        self.auto_adjust_columns(ws, headers)
        
        # Aplicar alineaciones específicas por columna
        # ID: centro
        ws.column_dimensions['A'].width = 10
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Fecha: centro
        ws.column_dimensions['B'].width = 20
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Usuario: izquierda
        ws.column_dimensions['C'].width = 20
        
        # Acción: izquierda
        ws.column_dimensions['D'].width = 15
        
        # Entidad: izquierda
        ws.column_dimensions['E'].width = 15
        
        # ID Entidad: centro
        ws.column_dimensions['F'].width = 12
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, min_col=6, max_col=6):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Descripción: izquierda, más ancha
        ws.column_dimensions['G'].width = 50
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

